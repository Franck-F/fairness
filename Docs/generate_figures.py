"""Generate Docs/figures_memoire.xlsx with all figures and tables for the AuditIQ memoire.

Run: python Docs/generate_figures.py
"""
from __future__ import annotations

import os
import random
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- Styles ----------
FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="1E3A8A")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="1E3A8A")
BODY_FONT = Font(name=FONT_NAME, size=11)
NOTE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")
HIGHLIGHT_FILL = PatternFill("solid", start_color="FEF08A")  # jaune
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Risk level fills (green -> red)
RISK_FILLS = {
    "min": PatternFill("solid", start_color="86EFAC"),
    "lim": PatternFill("solid", start_color="FDE68A"),
    "haut": PatternFill("solid", start_color="FCA5A5"),
    "int": PatternFill("solid", start_color="7F1D1D"),
}

# Complexity fills
COMPLEX_FILLS = {
    "vert": PatternFill("solid", start_color="BBF7D0"),
    "orange": PatternFill("solid", start_color="FED7AA"),
    "rouge": PatternFill("solid", start_color="FCA5A5"),
}


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, row_start: int, row_end: int, n_cols: int) -> None:
    for r in range(row_start, row_end + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font is None or cell.font.name != FONT_NAME:
                cell.font = BODY_FONT
            cell.alignment = LEFT if c > 1 else LEFT
            cell.border = BORDER


def autosize(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def set_title(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 24


def set_note(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT


# ---------- Build workbook ----------
wb = Workbook()
wb.remove(wb.active)

# ============ README ============
ws = wb.create_sheet("README")
set_title(ws, "AuditIQ - Figures et tableaux du memoire")
ws["A3"] = "Source des donnees sondage"
ws["A3"].font = Font(name=FONT_NAME, size=11, bold=True)
ws["A4"] = "Sondage IA et Ethique dans les PME - Enquete 2025 - N=34 repondants"
ws["A5"] = "Methodologie : enquete en ligne diffusee aupres de dirigeants et responsables IT/IA de PME francophones."
ws["A4"].font = BODY_FONT
ws["A5"].font = BODY_FONT

ws["A7"] = "Index des onglets"
ws["A7"].font = Font(name=FONT_NAME, size=11, bold=True)

index_rows = [
    ("Onglet", "Contenu", "Type"),
    ("F1_Repartition_secteurs", "Figure 1 - Repartition sectorielle des repondants", "Camembert"),
    ("F2_Connaissance_biais", "Figure 2 - Connaissance des biais algorithmiques (Q4)", "Barres horizontales"),
    ("F3_Budget_audit", "Figure 3 - Budget annuel audit ethique IA (Q7)", "Histogramme + reference"),
    ("F4_Pyramide_AI_Act", "Figure 4 - Pyramide des risques AI Act", "Tableau 4 niveaux"),
    ("F5_Metriques_fairness", "Figure 5 - Comparatif des metriques de fairness", "Tableau 5 col."),
    ("F6_Comparatif_outils", "Figure 6 - Comparatif des outils de detection des biais", "Tableau 7 col."),
    ("F9_Interet_outil", "Figure 9 - Interet pour un outil de detection (Q6)", "Barres horizontales"),
    ("F11_Architecture", "Figure 11 - Architecture technique AuditIQ", "Tableau couches"),
    ("F17_Triple_fosse", "Figure 17 - Schema de synthese du triple fosse", "Tableau 3 col."),
    ("Donnees_sondage_brutes", "Donnees individuelles simulees des 34 repondants", "Donnees brutes"),
]
for i, row in enumerate(index_rows, start=9):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)
style_header(ws, 9, 3)
style_body(ws, 10, 9 + len(index_rows) - 1, 3)

ws["A21"] = "Mode d'emploi"
ws["A21"].font = Font(name=FONT_NAME, size=11, bold=True)
ws["A22"] = "1. Chaque onglet contient les donnees brutes ainsi que le graphique correspondant quand pertinent."
ws["A23"] = "2. Modifiez les pourcentages ou libelles : les graphiques se mettent a jour automatiquement."
ws["A24"] = "3. L'onglet Donnees_sondage_brutes permet de recalculer toute statistique croisee (tableaux croises dynamiques)."
ws["A25"] = "4. Les couleurs respectent la charte du memoire (bleu nuit #1E3A8A pour les en-tetes)."
for r in range(22, 26):
    ws.cell(row=r, column=1).font = BODY_FONT

set_note(ws, 27, 1, "Genere automatiquement via Docs/generate_figures.py - ne pas editer sans regenerer.")
autosize(ws, {"A": 32, "B": 55, "C": 22})

# ============ F1_Repartition_secteurs ============
ws = wb.create_sheet("F1_Repartition_secteurs")
set_title(ws, "Figure 1 - Repartition sectorielle des repondants (N=34)")
headers = ["Secteur", "Part (%)"]
data = [
    ("Technologies / Informatique", 32.4),
    ("Finance / Assurance", 20.6),
    ("Autre", 20.6),
    ("Sante / Medical", 11.8),
    ("Services aux entreprises", 8.8),
    ("Industrie", 5.9),
]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 2)
for i, (lab, val) in enumerate(data, start=4):
    ws.cell(row=i, column=1, value=lab)
    c = ws.cell(row=i, column=2, value=val / 100)
    c.number_format = "0.0%"
style_body(ws, 4, 3 + len(data), 2)
# Total via formula
total_row = 3 + len(data) + 1
ws.cell(row=total_row, column=1, value="Total")
ws.cell(row=total_row, column=1).font = Font(name=FONT_NAME, size=11, bold=True)
tc = ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{3+len(data)})")
tc.number_format = "0.0%"
tc.font = Font(name=FONT_NAME, size=11, bold=True)

pie = PieChart()
labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(data))
vals = Reference(ws, min_col=2, min_row=3, max_row=3 + len(data))
pie.add_data(vals, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Repartition sectorielle (%)"
pie.dataLabels = DataLabelList(showPercent=True)
pie.height = 9
pie.width = 15
ws.add_chart(pie, "D3")

set_note(ws, total_row + 2, 1, "Source : Sondage IA et Ethique dans les PME - Enquete 2025, Q1 (secteur principal).")
autosize(ws, {"A": 32, "B": 12})

# ============ F2_Connaissance_biais ============
ws = wb.create_sheet("F2_Connaissance_biais")
set_title(ws, "Figure 2 - Connaissance des biais algorithmiques (Q4)")
for j, h in enumerate(["Niveau de connaissance", "Part (%)"], 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 2)
f2 = [
    ("Non, jamais", 38.2, "FCA5A5"),  # rouge
    ("Vaguement", 20.6, "FED7AA"),    # orange clair
    ("J'en ai entendu parler", 17.6, "FDE68A"),  # jaune
    ("Oui, je connais bien", 23.5, "86EFAC"),    # vert
]
for i, (lab, val, color) in enumerate(f2, start=4):
    ws.cell(row=i, column=1, value=lab)
    c = ws.cell(row=i, column=2, value=val / 100)
    c.number_format = "0.0%"
    ws.cell(row=i, column=1).fill = PatternFill("solid", start_color=color)
style_body(ws, 4, 3 + len(f2), 2)

bar = BarChart()
bar.type = "bar"
bar.style = 11
bar.title = "Connaissance des biais algorithmiques"
bar.y_axis.title = "Part des repondants"
bar.x_axis.title = ""
vals = Reference(ws, min_col=2, min_row=3, max_row=3 + len(f2))
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(f2))
bar.add_data(vals, titles_from_data=True)
bar.set_categories(cats)
bar.height = 8
bar.width = 16
ws.add_chart(bar, "D3")

set_note(ws, 10, 1, "Source : Sondage 2025, Q4 - Connaissez-vous la notion de biais algorithmique ?")
autosize(ws, {"A": 30, "B": 12})

# ============ F3_Budget_audit ============
ws = wb.create_sheet("F3_Budget_audit")
set_title(ws, "Figure 3 - Budget annuel pour audit ethique IA (Q7)")
for j, h in enumerate(["Tranche de budget", "Part (%)"], 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 2)
f3 = [
    ("0 EUR", 35.3),
    ("< 500 EUR", 17.6),
    ("500 - 2000 EUR", 29.4),
    ("> 2000 EUR", 5.9),
    ("Autre / NSP", 11.8),
]
for i, (lab, val) in enumerate(f3, start=4):
    ws.cell(row=i, column=1, value=lab)
    c = ws.cell(row=i, column=2, value=val / 100)
    c.number_format = "0.0%"
style_body(ws, 4, 3 + len(f3), 2)

# Reference info
ws.cell(row=11, column=1, value="Cout minimum audit externe estime (EUR)").font = Font(name=FONT_NAME, size=11, bold=True)
ws.cell(row=11, column=2, value=5000).number_format = "#,##0 \"EUR\""
set_note(ws, 12, 1, "Note : aucun repondant ne dispose d'un budget proche du cout minimum d'un audit externe (~5 000 EUR).")

bar = BarChart()
bar.type = "col"
bar.style = 12
bar.title = "Budget annuel declare (%)"
vals = Reference(ws, min_col=2, min_row=3, max_row=3 + len(f3))
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(f3))
bar.add_data(vals, titles_from_data=True)
bar.set_categories(cats)
bar.height = 8
bar.width = 16
ws.add_chart(bar, "D3")

set_note(ws, 14, 1, "Source : Sondage 2025, Q7 - Budget annuel alloue a l'audit ethique IA.")
autosize(ws, {"A": 32, "B": 14})

# ============ F4_Pyramide_AI_Act ============
ws = wb.create_sheet("F4_Pyramide_AI_Act")
set_title(ws, "Figure 4 - Pyramide de classification des risques de l'AI Act")
headers = ["Niveau", "Description", "Obligations", "Exemples PME"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 4)

f4 = [
    ("Risque minimal", "Aucune exigence specifique", "Aucune obligation", "Filtres anti-spam, jeux video", "min"),
    ("Risque limite", "Systemes interagissant avec utilisateurs", "Transparence (art. 50)", "Chatbots, deepfakes", "lim"),
    ("Haut risque", "Decisions a fort impact sur les droits", "Articles 9-15 complets", "Recrutement, scoring credit, acces services", "haut"),
    ("Pratiques interdites", "Atteintes aux valeurs fondamentales UE", "Article 5 - interdiction totale", "Notation sociale, manipulation subliminale", "int"),
]
for i, (lvl, desc, obl, ex, key) in enumerate(f4, start=4):
    ws.cell(row=i, column=1, value=lvl)
    ws.cell(row=i, column=2, value=desc)
    ws.cell(row=i, column=3, value=obl)
    ws.cell(row=i, column=4, value=ex)
    for c in range(1, 5):
        cell = ws.cell(row=i, column=c)
        cell.fill = RISK_FILLS[key]
        if key == "int":
            cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        else:
            cell.font = Font(name=FONT_NAME, size=11, bold=(c == 1))
style_body(ws, 4, 3 + len(f4), 4)
# re-apply fill after style_body (style_body sets borders, not fill)

set_note(ws, 9, 1, "Source : Reglement (UE) 2024/1689 sur l'IA (AI Act), articles 5, 6, 50 et annexe III.")
autosize(ws, {"A": 22, "B": 38, "C": 32, "D": 40})
for r in range(4, 8):
    ws.row_dimensions[r].height = 36

# ============ F5_Metriques_fairness ============
ws = wb.create_sheet("F5_Metriques_fairness")
set_title(ws, "Figure 5 - Comparatif des metriques de fairness")
headers = ["Metrique", "Definition simplifiee", "Ce qu'elle mesure", "Quand l'utiliser", "Limite principale"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 5)

f5 = [
    ("Demographic Parity", "Egalite du taux de resultats positifs entre groupes", "Parite statistique brute", "Quand l'objectif est l'egalite de representation", "Ignore la performance individuelle", "vert"),
    ("Equal Opportunity", "Egalite du taux de vrais positifs", "Egalite des chances pour les qualifies", "Quand on veut eviter de penaliser les groupes meritants", "Ne capte pas les faux positifs", "orange"),
    ("Equalized Odds", "Egalite TVP et TFP", "Cotes egalisees strictes", "Pour les contextes a haute criticite", "Souvent impossible a atteindre", "rouge"),
    ("Regle des 4/5", "Ratio min/max >= 0,8", "Seuil pragmatique juridique", "Quand on cherche un critere legalement defendable", "Moins formellement rigoureuse", "vert"),
]
for i, row in enumerate(f5, start=4):
    key = row[-1]
    for c, val in enumerate(row[:-1], start=1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.fill = COMPLEX_FILLS[key]
        cell.font = Font(name=FONT_NAME, size=11, bold=(c == 1))
style_body(ws, 4, 3 + len(f5), 5)

set_note(ws, 9, 1, "Code couleur : vert = simple a calculer, orange = intermediaire, rouge = avance.")
set_note(ws, 10, 1, "Source : Barocas, Hardt & Narayanan (2023), Fairness and Machine Learning ; guide EEOC (Four-Fifths rule).")
autosize(ws, {"A": 22, "B": 42, "C": 32, "D": 45, "E": 38})
for r in range(4, 8):
    ws.row_dimensions[r].height = 42

# ============ F6_Comparatif_outils ============
ws = wb.create_sheet("F6_Comparatif_outils")
set_title(ws, "Figure 6 - Comparatif des outils de detection des biais")
headers = ["Outil", "Developpeur", "Open Source", "Interface graphique", "Lien AI Act", "Cout", "Adapte PME"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 7)
f6 = [
    ("Fairlearn", "Microsoft / Linux Foundation", "Oui", "Non", "Non", "Gratuit", "Non"),
    ("AIF360", "IBM Research", "Oui", "Non", "Non", "Gratuit", "Non"),
    ("FAT Forensics", "U. Bristol", "Oui", "Non", "Non", "Gratuit", "Non"),
    ("LangBiTe", "UOC / U. Luxembourg", "Oui", "Non", "Partiel", "Gratuit", "Non"),
    ("Algorithm Audit", "Algorithm Audit NL", "Oui", "Oui", "Non", "Gratuit", "Partiel"),
    ("Fiddler AI", "Fiddler", "Non", "Oui", "Partiel", "Eleve", "Non"),
    ("TruEra (Snowflake)", "Snowflake", "Non", "Oui", "Partiel", "Eleve", "Non"),
    ("AuditIQ", "Franck F.", "Oui", "Oui", "Oui", "Gratuit", "Oui"),
]
for i, row in enumerate(f6, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        if row[0] == "AuditIQ":
            cell.fill = HIGHLIGHT_FILL
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
style_body(ws, 4, 3 + len(f6), 7)

set_note(ws, 13, 1, "Sources : documentations officielles de chaque outil (2024-2025). AuditIQ en surbrillance = contribution de ce memoire.")
autosize(ws, {"A": 22, "B": 30, "C": 14, "D": 20, "E": 14, "F": 12, "G": 14})

# ============ F9_Interet_outil ============
ws = wb.create_sheet("F9_Interet_outil")
set_title(ws, "Figure 9 - Interet pour un outil de detection (Q6)")
for j, h in enumerate(["Niveau d'interet", "Part (%)"], 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 2)
f9 = [
    ("Tres interesse, besoin urgent", 20.6, "BBF7D0"),
    ("Interesse si simple", 44.1, "BBF7D0"),
    ("Moyennement interesse", 23.5, "BBF7D0"),
    ("Pas interesse", 11.8, "FCA5A5"),
]
for i, (lab, val, color) in enumerate(f9, start=4):
    ws.cell(row=i, column=1, value=lab)
    c = ws.cell(row=i, column=2, value=val / 100)
    c.number_format = "0.0%"
    ws.cell(row=i, column=1).fill = PatternFill("solid", start_color=color)
style_body(ws, 4, 3 + len(f9), 2)

ws.cell(row=9, column=1, value="Total interesses (3 premieres lignes)").font = Font(name=FONT_NAME, size=11, bold=True)
c = ws.cell(row=9, column=2, value="=SUM(B4:B6)")
c.number_format = "0.0%"
c.font = Font(name=FONT_NAME, size=11, bold=True)

bar = BarChart()
bar.type = "bar"
bar.style = 11
bar.title = "Interet pour un outil de detection (%)"
vals = Reference(ws, min_col=2, min_row=3, max_row=3 + len(f9))
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(f9))
bar.add_data(vals, titles_from_data=True)
bar.set_categories(cats)
bar.height = 8
bar.width = 16
ws.add_chart(bar, "D3")

set_note(ws, 11, 1, "Source : Sondage 2025, Q6. Regroupement visuel : 88,2 % des repondants expriment un interet (vert).")
autosize(ws, {"A": 32, "B": 12})

# ============ F11_Architecture ============
ws = wb.create_sheet("F11_Architecture")
set_title(ws, "Figure 11 - Architecture technique d'AuditIQ")
headers = ["Couche", "Technologie", "Role", "Fournisseur"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 4)
f11 = [
    ("Frontend", "Next.js 14 + Tailwind + Shadcn/UI", "Interface utilisateur", "Vercel / Netlify"),
    ("Backend API", "FastAPI + Python 3.10", "Logique metier", "Render"),
    ("Base de donnees", "PostgreSQL", "Persistance", "Supabase"),
    ("Moteur fairness", "Fairlearn + Scikit-learn", "Calculs metriques", "Open source"),
    ("Module non supervise", "Scikit-learn KMeans + SciPy chi2", "Detection sans labels", "Open source"),
    ("Module LLM", "httpx + banque prompts", "Audit chatbot", "Custom"),
    ("Recommandations IA", "Google Gemini", "Generation NL", "Google"),
    ("Rapports", "ReportLab + OpenPyXL", "PDF / Excel + ancrage juridique FR", "Open source"),
]
for i, row in enumerate(f11, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        if c == 1:
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
style_body(ws, 4, 3 + len(f11), 4)
set_note(ws, 13, 1, "Source : architecture AuditIQ v1.0 (2025). Voir Docs/README.md du projet.")
autosize(ws, {"A": 24, "B": 38, "C": 38, "D": 20})

# ============ F17_Triple_fosse ============
ws = wb.create_sheet("F17_Triple_fosse")
set_title(ws, "Figure 17 - Schema de synthese du triple fosse")
headers = ["Fosse", "Constat sondage", "Reponse AuditIQ"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, 3)
f17 = [
    ("Cognitif", "38 % ne connaissent pas les biais", "Explications langage naturel + glossaire FR"),
    ("Technique", "Outils existants requierent du code", "Interface graphique guidee (wizard)"),
    ("Reglementaire", "Pas de lien metriques <-> AI Act", "Rapports avec ancrage Code du travail / CNIL"),
]
for i, row in enumerate(f17, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        if c == 1:
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.fill = PatternFill("solid", start_color="DBEAFE")
style_body(ws, 4, 3 + len(f17), 3)
for r in range(4, 7):
    ws.row_dimensions[r].height = 30
set_note(ws, 8, 1, "Source : synthese memoire AuditIQ, chapitre de conclusion.")
autosize(ws, {"A": 20, "B": 42, "C": 52})

# ============ Donnees_sondage_brutes ============
ws = wb.create_sheet("Donnees_sondage_brutes")
set_title(ws, "Donnees brutes simulees du sondage (N=34)")
headers = ["ID", "Secteur", "Taille_entreprise", "Usages_IA", "Connaissance_biais", "Perception_risque", "Interet_outil", "Budget_audit"]
for j, h in enumerate(headers, 1):
    ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(headers))

# Build deterministic dataset matching published percentages
random.seed(42)

def counts_from_pct(pcts, n=34):
    """Convert percentages to counts summing to n using largest remainder."""
    raw = [p / 100 * n for p in pcts]
    floor = [int(x) for x in raw]
    rem = [r - int(r) for r in raw]
    diff = n - sum(floor)
    order = sorted(range(len(rem)), key=lambda i: -rem[i])
    for i in range(diff):
        floor[order[i]] += 1
    return floor

secteurs = ["Technologies/Informatique", "Finance/Assurance", "Autre", "Sante/Medical", "Services aux entreprises", "Industrie"]
sect_pct = [32.4, 20.6, 20.6, 11.8, 8.8, 5.9]
sect_counts = counts_from_pct(sect_pct)

connaiss = ["Non, jamais", "Vaguement", "J'en ai entendu parler", "Oui, je connais bien"]
conn_pct = [38.2, 20.6, 17.6, 23.5]
conn_counts = counts_from_pct(conn_pct)

interet = ["Tres interesse, besoin urgent", "Interesse si simple", "Moyennement interesse", "Pas interesse"]
int_pct = [20.6, 44.1, 23.5, 11.8]
int_counts = counts_from_pct(int_pct)

budget = ["0 EUR", "< 500 EUR", "500-2000 EUR", "> 2000 EUR", "Autre"]
bud_pct = [35.3, 17.6, 29.4, 5.9, 11.8]
bud_counts = counts_from_pct(bud_pct)

def expand(labels, counts):
    out = []
    for lab, c in zip(labels, counts):
        out.extend([lab] * c)
    return out

col_sect = expand(secteurs, sect_counts)
col_conn = expand(connaiss, conn_counts)
col_int = expand(interet, int_counts)
col_bud = expand(budget, bud_counts)

random.shuffle(col_conn)
random.shuffle(col_int)
random.shuffle(col_bud)

tailles = ["< 10", "10-49", "50-249"]
usages_choices = ["ChatGPT/LLM", "Automatisation workflow", "Analyse donnees", "Recrutement", "Aucun"]
perceptions = ["Faible", "Moyenne", "Elevee"]

for i in range(34):
    row = 4 + i
    ws.cell(row=row, column=1, value=i + 1)
    ws.cell(row=row, column=2, value=col_sect[i])
    ws.cell(row=row, column=3, value=random.choice(tailles))
    ws.cell(row=row, column=4, value=random.choice(usages_choices))
    ws.cell(row=row, column=5, value=col_conn[i])
    ws.cell(row=row, column=6, value=random.choice(perceptions))
    ws.cell(row=row, column=7, value=col_int[i])
    ws.cell(row=row, column=8, value=col_bud[i])

style_body(ws, 4, 37, len(headers))

ws.cell(row=39, column=1, value="Total repondants").font = Font(name=FONT_NAME, size=11, bold=True)
c = ws.cell(row=39, column=2, value="=COUNTA(B4:B37)")
c.font = Font(name=FONT_NAME, size=11, bold=True)

set_note(ws, 41, 1, "Donnees individuelles simulees pour reproduire les distributions publiees (methode plus grand reste).")
set_note(ws, 42, 1, "Source : generees par Docs/generate_figures.py a partir du sondage 2025 (N=34).")
autosize(ws, {"A": 6, "B": 28, "C": 16, "D": 24, "E": 24, "F": 18, "G": 28, "H": 16})

# ---------- Save ----------
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "figures_memoire.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print("Sheets:", wb.sheetnames)
